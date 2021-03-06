#!/usr/bin/env python3.7

import argparse
import requests
from bs4 import BeautifulSoup, Tag
import json
import os
import platform
import re
import time

from openpyxl import load_workbook, Workbook

from replit_keep_alive import replit_keep_alive

#todo list:
'''
- add subito categories
- fix adding existing results
'''




def load_excel_db():
    global wb
    try:
        wb = load_workbook(filename=global_excel_db)
    except:
        # create excel db if not existing
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Searches"
        sheet = wb["Searches"]
        column_names = 'Active,Search Name,MinPrice,MaxPrice,SearchArea,Category,Keywords,KeywordsEclude,Only in Title,Only Can Post,Website,URL Search'.split(
            ',')
        for cl in range(len(column_names)):
            sheet.cell(row=1, column=cl + 1, value=column_names[cl])
        wb.save(filename=global_excel_db)

        sheet.create_sheet("Results")
        sheet = wb["Results"]
        column_names_results = 'Search Name, Title, Price, Location, Link, Search URL'.split(',')
        for clr in range(len(column_names)):
            sheet.cell(row=1, column=clr + 1, value=column_names[clr])
        wb.save(filename=global_excel_db)


def extract_column_names_and_numbers_to_dict(worksheet):
    col_names_number = {}
    current_val = 1
    for cnm in worksheet.iter_cols(1, worksheet.max_column):
        col_names_number[cnm[0].value] = current_val
        current_val += 1
    return col_names_number


def check_and_fill_empty_url_in_excel_db():
    global col_names_number, ws_row_number

    def fill_search_url_in_excel_db():
        global min_price, max_price
        # example1='https://www.subito.it/annunci-italia/vendita/arredamento-casalinghi/?q=herman+miller&order=priceasc&ps=100&pe=350'
        url_root = 'https://www.subito.it/annunci-'
        search_area_dic = {
            'italia': 'italia',
            'EM': 'emilia-romagna',
            'EM vicino': 'emilia-romagna-vicino',
            'RE': '/reggio-emilia',
            'MO': '/modena',
            'PR': '/parma'
        }
        categoria_dic = {
            'Informatica': 'informatica',
            'Elettronica': 'elettronica',
            'Audio Video': 'audio-video',
            'Fotografia': 'fotografia',
            'Telefonia': 'telefonia',
            'Per la casa e persona': 'casa-e-persona',
            'Arredamento e Casalinghi': 'arredamento-casalinghi',
            'Giardino e Fai da te': 'giardino-fai-da-te',
            'Abbigliamento e Accessori': 'abbigliamento-accessori',
            'Sport e Hobby': 'sport-hobby',
            'Libri e Riviste': 'libri-riviste',
            'Strumenti Musicali': 'strumenti-musicali',
            'Biciclette': 'biciclette',
            '': '',
            '': '',
            '': '',
        }
        search_area = ws.cell(row=ws_row_number, column=col_names_number['SearchArea']).value
        ##because provinces need region area and province sub areas in URL
        if search_area in ['RE', 'MO', 'PR']:
            sub_search_area = search_area_dic[search_area]
            search_area = search_area_dic['EM']

        else:
            search_area = search_area_dic[search_area]
            sub_search_area = ''
        category = ws.cell(row=ws_row_number, column=col_names_number['Category']).value
        category = categoria_dic[category]
        keywords = ws.cell(row=ws_row_number, column=col_names_number['Keywords']).value
        keywords = '+'.join([x.strip().lower() for x in keywords.split(' ')])
        print()
        min_price = ws.cell(row=ws_row_number, column=col_names_number['MinPrice']).value or None
        max_price = ws.cell(row=ws_row_number, column=col_names_number['MaxPrice']).value or None
        only_in_title = True if ws.cell(row=ws_row_number, column=col_names_number['Only in Title']).value else False
        only_can_post = True if ws.cell(row=ws_row_number, column=col_names_number['Only Can Post']).value else False
        search_url = url_root + search_area + '/vendita/' + category + sub_search_area + '/?q=' + keywords + '&ps=' + str(
            min_price) + '&pe=' + str(max_price)
        ws.cell(row=ws_row_number, column=col_names_number['URL Search']).value = search_url
        wb.save(filename=global_excel_db)

        #after adding url, searches for query
        query_name = ws.cell(row=ws_row_number, column=col_names_number['Search Name']).value
        run_query(url=search_url, name=query_name, notify=True)

    # Create a dictionary of column names
    col_names_number = extract_column_names_and_numbers_to_dict(worksheet=ws)
    url_row_number = col_names_number['URL Search']
    ##find columns that have no url
    for ws_row_number in range(1, ws.max_row + 1):
        if not ws.cell(row=ws_row_number, column=url_row_number).value:
            fill_search_url_in_excel_db()


def parser_set_up():
    parser = argparse.ArgumentParser()
    parser.add_argument("--add", dest='name', help="name of new tracking to be added")
    parser.add_argument("--url", help="url for your new tracking's search query")
    parser.add_argument("--delete", help="name of the search you want to delete")
    parser.add_argument('--refresh', '-r', dest='refresh', action='store_true', help="refresh search results once")
    parser.set_defaults(refresh=False)
    parser.add_argument('--daemon', '-d', dest='daemon', action='store_true',
                        help="keep refreshing search results forever (default delay 120 seconds)")
    parser.set_defaults(daemon=False)
    parser.add_argument('--delay', dest='delay', help="delay for the daemon option (in seconds)")
    parser.set_defaults(delay=120)
    parser.add_argument('--list', dest='list', action='store_true', help="print a list of current trackings")
    parser.set_defaults(list=False)
    parser.add_argument('--short_list', dest='short_list', action='store_true', help="print a more compact list")
    parser.set_defaults(short_list=False)
    parser.add_argument('--tgoff', dest='tgoff', action='store_true', help="turn off telegram messages")
    parser.set_defaults(tgoff=False)
    parser.add_argument('--notifyoff', dest='win_notifyoff', action='store_true', help="turn off windows notifications")
    parser.set_defaults(win_notifyoff=False)
    parser.add_argument('--addtoken', dest='token', help="telegram setup: add bot API token")
    parser.add_argument('--addchatid', dest='chatid', help="telegram setup: add bot chat id")
    # my parsers
    parser.add_argument("--only_with_price", dest='only_with_price', help="show ads without price: default True")
    parser.set_defaults(only_with_price=True)
    parser.add_argument("--min_price", dest='min_price', help="min price for captured ads; default 0")
    parser.set_defaults(min_price=0)
    parser.add_argument("--max_price", dest='max_price', help="max price for captured ads; default 10.000")
    parser.set_defaults(min_price=10000)
    parser.add_argument("--location", dest='location', help="ads geographical location")
    parser.set_defaults(location='Reggio Emilia')
    parser.add_argument("--keyword_in_title", dest='keyword_in_title',
                        help="show only ads with keyword in the title: default False")
    parser.set_defaults(keyword_in_title=False)
    parser.add_argument("--first_notify", dest='first_notify',
                        help="notifies all results upon first search: default False")
    parser.set_defaults(first_notify=False)
    args = parser.parse_args()
    return args


def data_from_db_to_dict(ws):
    db_parent_dict = {}
    col_title_numbers = extract_column_names_and_numbers_to_dict(worksheet=ws)
    for row in range(2, ws.max_row + 1):
        # dict_parent_name= str(row) + ' - ' +  col_title_numbers['Search Name']
        dict_parent_name= str(row) + ' - ' +  ws.cell(row=1, column=col_title_numbers['Search Name']).value
        db_child_dict={}
        for col in range(1, ws.max_column+1):
            cell_title = ws.cell(row=1, column=col).value
            cell_content = ws.cell(row=row, column=col).value
            # db_child_dict{cell_title : cell_content}
            db_child_dict[cell_title] = cell_content
        db_parent_dict[dict_parent_name] = db_child_dict
    return db_parent_dict


def load_api_credentials():
    global apiCredentials
    global telegramApiFile
    if not os.path.isfile(telegramApiFile):
        return

    with open(telegramApiFile) as file:
        apiCredentials = json.load(file)


def save_results_to_excel_db(name, url, link, title, price, location):
    global wb
    global global_excel_db
    wb.save(global_excel_db)
    ws = wb["Results"]
    col_names_numbers = extract_column_names_and_numbers_to_dict(worksheet=ws)

    all_results = data_from_db_to_dict(ws)
    all_results_urls = []
    for res in all_results.values():
        all_results_urls.append(res['Link'])
    is_new_result = False
    if link not in all_results_urls:
        new_row = ws.max_row+1
        ws.cell(row=new_row, column=col_names_numbers['Search Name']).value = name
        ws.cell(row=new_row, column=col_names_numbers['Title']).value = title
        ws.cell(row=new_row, column=col_names_numbers['Price']).value = price
        ws.cell(row=new_row, column=col_names_numbers['Location']).value = location
        ws.cell(row=new_row, column=col_names_numbers['Link']).value = link
        ws.cell(row=new_row, column=col_names_numbers['Search URL']).value = url
        wb.save(global_excel_db)
        is_new_result = True

    ws = wb['Searches']
    return is_new_result


def run_query(url, name, notify):
    print("running query (\"{}\" - {})...".format(name, url))
    global queries
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'html.parser')

    product_list_items = soup.find_all('div', class_=re.compile(r'item-key-data'))
    msg = []

    for product in product_list_items:
        title = product.find('h2').string

        try:
            price = product.find('p', class_=re.compile(r'price')).contents[0]
            # check if the span tag exists
            price_soup = BeautifulSoup(price, 'html.parser')
            if type(price_soup) == Tag:
                continue
            # at the moment (20.5.2021) the price is under the 'p' tag with 'span' inside if shipping available

        except:
            price = "Unknown price"
        link = product.parent.parent.parent.parent.get('href')

        location = product.find('span', re.compile(r'town')).string + product.find('span', re.compile(r'city')).string

        # if not queries.get(name):  # insert the new search
        #     queries[name] = {url: {link: {'title': title, 'price': price, 'location': location}}}
        #     print("\nNew search added:", name)
        #     print("Adding result:", title, "-", price, "-", location)
        # else:  # add search results to dictionary
        #     if not queries.get(name).get(url).get(link):  # found a new element
        #         tmp = "New element found for " + name + ": " + title + " @ " + price + " - " + location + " --> " + link + '\n'
        #         msg.append(tmp)
        #         queries[name][url][link] = {'title': title, 'price': price, 'location': location}

        is_new_result = save_results_to_excel_db(name, url, link, title, price, location)
        # tmp = "New element found for " + name + ": " + title + " @ " + price + " - " + location + " --> " + link + '\n'
        if is_new_result:
            msg.append("New element found for " + name + ": " + title + " @ " + price + " - " + location + " --> " + link + '\n')

    if len(msg) > 0:
        if notify:
            # # Windows only: send notification
            # if not args.win_notifyoff and platform.system() == "Windows":
            #     global toaster
            #     toaster.show_toast("New announcements", "Query: " + name)
            if is_telegram_active():
                send_telegram_messages(msg)
            print("\n".join(msg))
            print('\n{} new elements have been found.'.format(len(msg)))
        # save_queries()
    else:
        print('All lists are already up to date.\n')
    # print("queries file saved: ", queries)


def refresh_search(notify=False):
    global db_dict
    try:
        for search in db_dict.values():
            run_query(url=search['URL Search'], name=search['Search Name'], notify=notify)
    except requests.exceptions.ConnectionError:
        print("***Connection error***")
    except requests.exceptions.Timeout:
        print("***Server timeout error***")
    except requests.exceptions.HTTPError:
        print("***HTTP error***")


def save_api_credentials():
    with open(telegramApiFile, 'w') as file:
        file.write(json.dumps(apiCredentials))


def is_telegram_active():
    global args
    return not args.tgoff and "chatid" in apiCredentials and "token" in apiCredentials


def send_telegram_messages(messages):
    for msg in messages:
        request_url = "https://api.telegram.org/bot" + apiCredentials["token"] + "/sendMessage?chat_id=" + \
                      apiCredentials["chatid"] + "&text=" + msg
        requests.get(request_url)

# KEEP SCRIPT RUNNING 24/7 WITH REPLIT
replit_keep_alive.keep_alive()

##NAME OF EXCEL DB
# global_excel_db = 'not_ready.xlsx'
global_excel_db = 'subito.xlsx'
args = parser_set_up()

telegramApiFile = "telegram_api_credentials"
apiCredentials = dict()
load_api_credentials()

# MAIN
if __name__ == '__main__':

    load_excel_db()
    wb = load_workbook(filename=global_excel_db)
    ws = wb["Searches"]

    check_and_fill_empty_url_in_excel_db()

    db_dict = data_from_db_to_dict(ws)

    # args = parser_set_up()

    # queries = dict()
    queries = db_dict

    # dbFile = "searches.tracked"

    # telegramApiFile = "telegram_api_credentials"
    # apiCredentials = dict()
    # load_api_credentials()


    # Telegram setup
    if args.token is not None and args.chatid is not None:
        apiCredentials["token"] = args.token
        apiCredentials["chatid"] = args.chatid
        save_api_credentials()

    first_notify = args.first_notify #commentare?
    if args.daemon:
        # notify = args.first_notify
        # notify = Fal se  # Don't flood with notifications the first time
        notify = True

        while True:
            refresh_search(notify)
            print(str(args.delay) + " seconds to next poll.")
            time.sleep(int(args.delay))
